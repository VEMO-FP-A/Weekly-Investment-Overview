"""
update_and_push.py
──────────────────
Script único para el Weekly Investment Dashboard de VEMO Treasury.

  1. Lee el Excel de inversiones (tab Yields)
  2. Actualiza el HTML del dashboard con los datos nuevos
  3. Sube todo a GitHub (VEMO-FP-A/Weekly-Investment-Overview)

Uso:
  python update_and_push.py                         # actualiza + sube
  python update_and_push.py --update-only           # solo actualiza HTML, no sube
  python update_and_push.py --push-only             # solo sube a GitHub, no toca HTML
  python update_and_push.py --dry-run               # preview sin escribir ni subir
  python update_and_push.py -m "Semana Jul 21"      # commit message personalizado
  python update_and_push.py --excel "otro.xlsx"     # archivo Excel distinto

Configuración del token de GitHub (una sola vez):
  Windows CMD:   set GITHUB_TOKEN=ghp_tu_token_aqui
  Windows PS:    $env:GITHUB_TOKEN="ghp_tu_token_aqui"
  Mac/Linux:     export GITHUB_TOKEN=ghp_tu_token_aqui

  Para que persista: agrégalo a Variables de entorno del sistema
  (Windows: System Properties > Environment Variables).

Requisitos:
  pip install openpyxl
"""

import os
import sys
import re
import json
import base64
import argparse
from datetime import datetime
from urllib import request, error

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════════════════
# PARTE 1 — EXTRACCIÓN DE DATOS DEL EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def extract(filepath: str) -> dict:
    """Extrae datos del tab Yields del Excel y devuelve un diccionario."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Yields"]

    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun"]

    # Mapeo de columnas: U(21)=nombre, W(23)=AVG/SUM, Y-AD(25-30)=Ene-Jun
    COL_NAME = 21
    COL_START = 25
    COL_END = 30
    COL_AVG = 23

    def read_row(row: int) -> dict:
        name = ws.cell(row=row, column=COL_NAME).value
        avg = ws.cell(row=row, column=COL_AVG).value
        values = []
        for c in range(COL_START, COL_END + 1):
            v = ws.cell(row=row, column=c).value
            values.append(v if v is not None else 0)
        return {"name": name, "avg": avg, "monthly": values}

    entities = []
    interest = {}
    cash = {}
    yields = {}

    for offset in range(0, 13):
        int_row = read_row(7 + offset)
        cash_row = read_row(24 + offset)
        yld_row = read_row(41 + offset)
        name = int_row["name"]
        if name is None:
            continue
        entities.append(name)
        interest[name] = {"total": int_row["avg"], "monthly": int_row["monthly"]}
        cash[name] = {"avg": cash_row["avg"], "monthly": cash_row["monthly"]}
        yields[name] = {"avg": yld_row["avg"], "monthly": yld_row["monthly"]}

    int_total = read_row(20)
    cash_total = read_row(37)
    yld_total = read_row(54)

    wb.close()

    return {
        "entities": entities,
        "months": months,
        "interest": interest,
        "cash": cash,
        "yields": yields,
        "total_interest_h1": int_total["avg"],
        "total_cash_avg": cash_total["avg"],
        "ytd_yield": yld_total["avg"],
        "total_interest_monthly": int_total["monthly"],
        "total_yield_monthly": yld_total["monthly"],
    }


def print_report(data: dict):
    """Imprime un resumen de los datos extraídos."""
    def fmt_money(v):
        if v is None: return "$0"
        sign = "-" if v < 0 else ""
        return f"{sign}${abs(v):,.0f}"

    def fmt_pct(v):
        if v is None: return "0.00%"
        return f"{v * 100:.2f}%"

    months = data["months"]
    print(f"\n{'INTERESES GANADOS H1 2026':^80}")
    header = f"  {'Empresa':<14}" + "".join(f"{m:>12}" for m in months) + f"{'Total H1':>14}"
    print(header)
    print("  " + "─" * 76)

    sorted_ents = sorted(
        data["entities"],
        key=lambda e: sum(data["interest"][e]["monthly"]),
        reverse=True,
    )
    for e in sorted_ents:
        vals = data["interest"][e]["monthly"]
        total = sum(vals)
        print(f"  {e:<14}" + "".join(f"{fmt_money(v):>12}" for v in vals) + f"{fmt_money(total):>14}")

    totals = data["total_interest_monthly"]
    grand = sum(totals)
    print("  " + "─" * 76)
    print(f"  {'TOTAL':<14}" + "".join(f"{fmt_money(v):>12}" for v in totals) + f"{fmt_money(grand):>14}")

    print(f"\n  Total Intereses H1:  {fmt_money(data['total_interest_h1'])}")
    print(f"  Cash Promedio H1:    {fmt_money(data['total_cash_avg'])}")
    print(f"  Yield YTD:           {fmt_pct(data['ytd_yield'])}")


# ═══════════════════════════════════════════════════════════════════════════════
# PARTE 1.5 — DATOS DE MERCADO VÍA APIs (FRED + Banxico)
# ═══════════════════════════════════════════════════════════════════════════════

# Proyecciones H2 2026 — actualizar manualmente con consenso de analistas
PROJ_TIIE = {
    "2026-08": 6.50, "2026-09": 6.50,
    "2026-10": 6.50, "2026-11": 6.50, "2026-12": 6.50,
}
PROJ_FED = {
    "2026-08": 3.75, "2026-09": 3.75,
    "2026-10": 3.75, "2026-11": 3.80, "2026-12": 3.80,
}
PROJ_FX = {
    "2026-08": 17.40, "2026-09": 17.50,
    "2026-10": 17.60, "2026-11": 17.70, "2026-12": 17.92,
}


def get_api_key(filenames):
    """Lee API key del primer archivo encontrado en SCRIPT_DIR."""
    for fn in filenames:
        path = os.path.join(SCRIPT_DIR, fn)
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                key = f.read().strip()
            if key:
                return key
    return None


def fetch_fred_series(series_id, fred_key, start="2022-01-01", freq="m", agg="eop"):
    """Descarga serie mensual de FRED. Devuelve {YYYY-MM: valor}."""
    end = datetime.now().strftime("%Y-%m-%d")
    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&api_key={fred_key}&file_type=json"
        f"&observation_start={start}&observation_end={end}"
        f"&frequency={freq}&aggregation_method={agg}"
    )
    req = request.Request(url, headers={"User-Agent": "VEMO-Treasury-Dashboard"})
    resp = request.urlopen(req, timeout=20)
    data = json.loads(resp.read())
    monthly = {}
    for obs in data.get("observations", []):
        val = obs["value"]
        if val and val != ".":
            ym = obs["date"][:7]
            monthly[ym] = round(float(val), 2)
    return monthly


def fetch_banxico_tiie(token, start="2022-01-01"):
    """Descarga TIIE 28d de Banxico SIE. Devuelve {YYYY-MM: último valor del mes}."""
    end = datetime.now().strftime("%Y-%m-%d")
    url = (
        f"https://www.banxico.org.mx/SieAPIRest/service/v1/"
        f"series/SF43783/datos/{start}/{end}?mediaType=json"
    )
    req = request.Request(url, headers={
        "Bmx-Token": token,
        "User-Agent": "VEMO-Treasury-Dashboard",
        "Accept": "application/json",
    })
    resp = request.urlopen(req, timeout=20)
    data = json.loads(resp.read())
    monthly = {}
    series = data.get("bmx", {}).get("series", [{}])[0]
    for dato in series.get("datos", []):
        fecha = dato["fecha"]  # dd/mm/yyyy
        val = dato.get("dato", "")
        if val and val not in ("N/E", ""):
            parts = fecha.split("/")
            ym = f"{parts[2]}-{parts[1]}"
            monthly[ym] = round(float(val.replace(",", "")), 2)
    return monthly


def build_market_data(tiie_raw, fed_raw, fx_raw):
    """Construye arrays de 60 elementos (Ene'22 - Dic'26) para las gráficas."""
    all_yms = []
    for y in range(2022, 2027):
        for m in range(1, 13):
            all_yms.append(f"{y}-{m:02d}")

    tiie_m = [tiie_raw.get(ym) for ym in all_yms]
    fed_m = [fed_raw.get(ym) for ym in all_yms]
    fx_m = [fx_raw.get(ym) for ym in all_yms]

    # Último mes donde las tres series tienen dato
    last_actual = -1
    for i in range(len(all_yms)):
        if tiie_m[i] is not None and fed_m[i] is not None and fx_m[i] is not None:
            last_actual = i

    # Nullificar meses posteriores
    for i in range(last_actual + 1, 60):
        tiie_m[i] = None
        fed_m[i] = None
        fx_m[i] = None

    proj_start = last_actual

    # Proyecciones (conectan desde el último dato real)
    tiie_proj = [None] * 60
    fed_proj = [None] * 60
    fx_proj = [None] * 60

    if proj_start >= 0:
        tiie_proj[proj_start] = tiie_m[proj_start]
        fed_proj[proj_start] = fed_m[proj_start]
        fx_proj[proj_start] = fx_m[proj_start]

    for i in range(proj_start + 1, 60):
        ym = all_yms[i]
        tiie_proj[i] = PROJ_TIIE.get(ym)
        fed_proj[i] = PROJ_FED.get(ym)
        fx_proj[i] = PROJ_FX.get(ym)

    return {
        "tiie_m": tiie_m, "fed_m": fed_m, "fx_m": fx_m,
        "tiie_proj": tiie_proj, "fed_proj": fed_proj, "fx_proj": fx_proj,
        "proj_start": proj_start,
        "tiie_26": tiie_m[48:60], "fed_26": fed_m[48:60], "fx_26": fx_m[48:60],
        "last_ym": all_yms[last_actual] if last_actual >= 0 else "N/A",
    }


def js_array(arr, decimals=2, per_line=12):
    """Convierte lista Python a string de array JS formateado."""
    parts = []
    for v in arr:
        if v is None:
            parts.append("null")
        else:
            parts.append(f"{v:.{decimals}f}")
    lines = []
    for i in range(0, len(parts), per_line):
        lines.append("  " + ",".join(parts[i:i + per_line]))
    return "[\n" + ",\n".join(lines) + "\n]"


def update_market_html(html, market):
    """Reemplaza arrays de datos de mercado en el HTML."""
    def replace_const(h, name, arr, dec=2):
        pattern = rf"const {name}=\[[\s\S]*?\];"
        repl = f"const {name}={js_array(arr, dec)};"
        return re.sub(pattern, repl, h)

    html = replace_const(html, "TIIE_M", market["tiie_m"])
    html = replace_const(html, "TIIE_PROJ", market["tiie_proj"])
    html = replace_const(html, "FED_M", market["fed_m"])
    html = replace_const(html, "FED_PROJ", market["fed_proj"])
    html = replace_const(html, "FX_M", market["fx_m"])
    html = replace_const(html, "FX_PROJ", market["fx_proj"])

    html = re.sub(
        r"const PROJ_START=\d+;.*",
        f'const PROJ_START={market["proj_start"]}; // último mes con data real',
        html,
    )

    html = replace_const(html, "TIIE_26", market["tiie_26"])
    html = replace_const(html, "FED_26", market["fed_26"])
    html = replace_const(html, "FX_26", market["fx_26"])

    return html


def fetch_market_data():
    """Descarga datos de mercado y construye los arrays para las gráficas."""
    fred_key = get_api_key(["fred_key.txt", "fred_key"])
    banxico_token = get_api_key(["banxico_token.txt", "banxico_token"])

    if not fred_key:
        print("  ⚠ No se encontró fred_key.txt")
    if not banxico_token:
        print("  ⚠ No se encontró banxico_token.txt")
    if not fred_key and not banxico_token:
        return None

    tiie_raw, fed_raw, fx_raw = {}, {}, {}

    if banxico_token:
        try:
            print("  Descargando TIIE 28d (Banxico)...", end=" ", flush=True)
            tiie_raw = fetch_banxico_tiie(banxico_token)
            print(f"✓ {len(tiie_raw)} meses")
        except Exception as e:
            print(f"✗ {e}")

    if fred_key:
        try:
            print("  Descargando Fed Funds Rate (FRED)...", end=" ", flush=True)
            fed_raw = fetch_fred_series("DFEDTARU", fred_key)
            print(f"✓ {len(fed_raw)} meses")
        except Exception as e:
            print(f"✗ {e}")
            try:
                print("    Fallback FEDFUNDS...", end=" ", flush=True)
                fed_raw = fetch_fred_series("FEDFUNDS", fred_key, agg="avg")
                print(f"✓ {len(fed_raw)} meses (tasa efectiva)")
            except Exception:
                print("✗")

        try:
            print("  Descargando FX MXN/USD (FRED)...", end=" ", flush=True)
            fx_raw = fetch_fred_series("DEXMXUS", fred_key)
            print(f"✓ {len(fx_raw)} meses")
        except Exception as e:
            print(f"✗ {e}")

    if not tiie_raw or not fed_raw or not fx_raw:
        missing = []
        if not tiie_raw: missing.append("TIIE")
        if not fed_raw: missing.append("Fed")
        if not fx_raw: missing.append("FX")
        print(f"  ⚠ Faltan: {', '.join(missing)} — se necesitan las 3 series")
        return None

    market = build_market_data(tiie_raw, fed_raw, fx_raw)
    print(f"  Último dato real: {market['last_ym']} (PROJ_START={market['proj_start']})")
    return market


# ═══════════════════════════════════════════════════════════════════════════════
# PARTE 2 — ACTUALIZACIÓN DEL HTML
# ═══════════════════════════════════════════════════════════════════════════════

def build_yield_int_js(data: dict) -> str:
    lines = ["const YIELD_INT={"]
    entities = data["entities"]
    for i, e in enumerate(entities):
        vals = data["interest"][e]["monthly"]
        arr = ",".join(f"{v:.2f}" for v in vals)
        comma = "," if i < len(entities) - 1 else ""
        lines.append(f"  {e}:[{arr}]{comma}")
    lines.append("};")
    return "\n".join(lines)


def build_yield_pct_js(data: dict) -> str:
    lines = [
        "// Pre-computed annualized yields from Excel (exact cached values from file)",
        "const YIELD_PCT={",
    ]
    for e in data["entities"]:
        vals = data["yields"][e]["monthly"]
        arr = ",".join(f"{v * 100:.2f}" for v in vals)
        lines.append(f"  {e}:[{arr}],")

    total_vals = data["total_yield_monthly"]
    arr = ",".join(f"{v * 100:.2f}" for v in total_vals)
    lines.append(f"  TOTAL:[{arr}]")
    lines.append("};")
    return "\n".join(lines)


def build_kpi_values(data: dict) -> dict:
    total_h1 = data["total_interest_h1"] or 0
    ytd_yield = data["ytd_yield"] or 0
    monthly_avg = total_h1 / 6 if total_h1 else 0
    return {
        "rendimiento_acumulado_display": f"${total_h1 / 1e6:.1f}M",
        "yield_ytd": f"{ytd_yield * 100:.2f}%",
        "rendimiento_mensual_display": f"${monthly_avg / 1e6:.2f}M",
    }


def update_html(html_path: str, data: dict) -> str:
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    # 1. Reemplazar YIELD_INT
    html = re.sub(r"const YIELD_INT=\{.*?\};", build_yield_int_js(data), html, flags=re.DOTALL)

    # 2. Reemplazar YIELD_PCT
    html = re.sub(
        r"// Pre-computed annualized yields.*?const YIELD_PCT=\{.*?\};",
        build_yield_pct_js(data), html, flags=re.DOTALL,
    )

    # 3. KPI cards
    kpi = build_kpi_values(data)
    html = re.sub(
        r'(Rendimiento Acumulado Real</div>\s*<div class="kpi-value"[^>]*>)\$[\d.]+M',
        rf'\g<1>{kpi["rendimiento_acumulado_display"]}', html,
    )
    html = re.sub(r"Yield YTD: [\d.]+%", f'Yield YTD: {kpi["yield_ytd"]}', html)

    return html


# ═══════════════════════════════════════════════════════════════════════════════
# PARTE 3 — PUSH A GITHUB
# ═══════════════════════════════════════════════════════════════════════════════

REPO_OWNER = "VEMO-FP-A"
REPO_NAME = "Weekly-Investment-Overview"
BRANCH = "main"
API_BASE = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}"

# (local_filename, repo_filename) — el dashboard se sube como index.html para GitHub Pages
FILES_TO_PUSH = [
    ("investment-dashboard.html", "index.html"),
    ("update_and_push.py", "update_and_push.py"),
]


CONFIG_FILE = os.path.join(SCRIPT_DIR, ".github_config")


def get_token() -> str:
    # 1. Leer del config guardado (runs anteriores)
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            token = f.read().strip()
        if token:
            return token
    # 2. Primera vez: leer del archivo 'token' o 'token.txt' y guardarlo en config
    token_file = None
    for name in ["token", "token.txt"]:
        candidate = os.path.join(SCRIPT_DIR, name)
        if os.path.exists(candidate):
            token_file = candidate
            break
    if token_file:
        with open(token_file, "r", encoding="utf-8") as f:
            token = f.read().strip()
        if token:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(token)
            print("  Token guardado en config (no se volverá a leer el archivo 'token')")
            return token
    # 3. Fallback: variable de entorno
    token = os.environ.get("GITHUB_TOKEN")
    if token:
        return token
    # 4. No se encontró
    print("\nError: No se encontró el token de GitHub.")
    print("Coloca un archivo llamado 'token' en la misma carpeta del script.")
    sys.exit(1)


def github_request(endpoint: str, token: str, method: str = "GET", data: dict = None):
    url = f"{API_BASE}{endpoint}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "VEMO-Treasury-Dashboard",
    }
    body = json.dumps(data).encode("utf-8") if data else None
    if body:
        headers["Content-Type"] = "application/json"

    req = request.Request(url, data=body, headers=headers, method=method)
    try:
        resp = request.urlopen(req)
        return json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as e:
        if e.code == 404 and method == "GET":
            return None
        err_body = e.read().decode("utf-8")
        print(f"  Error GitHub API ({e.code}): {err_body[:200]}")
        raise


def push_file(local_path: str, repo_path: str, token: str, message: str) -> bool:
    with open(local_path, "rb") as f:
        content = base64.b64encode(f.read()).decode("utf-8")

    result = github_request(f"/contents/{repo_path}?ref={BRANCH}", token)
    sha = result["sha"] if result else None

    payload = {"message": message, "content": content, "branch": BRANCH}
    if sha:
        payload["sha"] = sha

    try:
        github_request(f"/contents/{repo_path}", token, method="PUT", data=payload)
        return True
    except Exception:
        return False


def push_to_github(commit_msg: str, dry_run: bool = False):
    token = get_token()

    print(f"\n── GitHub: {REPO_OWNER}/{REPO_NAME} ──")
    try:
        repo_info = github_request("", token)
        if repo_info is None:
            print("Error: No se pudo acceder al repositorio.")
            sys.exit(1)
        print(f"  Conectado · Branch: {BRANCH} · {repo_info.get('visibility', '?')}")
    except Exception as e:
        print(f"Error de conexión: {e}")
        sys.exit(1)

    if dry_run:
        print("  [DRY RUN] No se subieron archivos.")
        for local_name, repo_name in FILES_TO_PUSH:
            path = os.path.join(SCRIPT_DIR, local_name)
            status = f"✓ {os.path.getsize(path):,}b" if os.path.exists(path) else "✗ no encontrado"
            label = f"{local_name} → {repo_name}" if local_name != repo_name else local_name
            print(f"    {label}: {status}")
        return

    print(f"  Commit: {commit_msg}\n")
    success = 0
    for local_name, repo_name in FILES_TO_PUSH:
        local_path = os.path.join(SCRIPT_DIR, local_name)
        if not os.path.exists(local_path):
            print(f"  ✗ {local_name} — no encontrado")
            continue
        size = os.path.getsize(local_path)
        label = f"{local_name} → {repo_name}" if local_name != repo_name else local_name
        print(f"  ↑ {label} ({size:,}b)...", end=" ", flush=True)
        if push_file(local_path, repo_name, token, commit_msg):
            print("✓")
            success += 1
        else:
            print("✗ error")

    print(f"\n  {success}/{len(FILES_TO_PUSH)} archivos subidos")
    print(f"  https://github.com/{REPO_OWNER}/{REPO_NAME}")
    print(f"  https://{REPO_OWNER.lower()}.github.io/{REPO_NAME}/")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="VEMO Treasury — Actualiza dashboard y sube a GitHub"
    )
    parser.add_argument("--excel", default=os.path.join(SCRIPT_DIR, "Investment analysis VEMO v2.xlsx"),
                        help="Ruta al archivo Excel")
    parser.add_argument("--html", default=os.path.join(SCRIPT_DIR, "investment-dashboard.html"),
                        help="Ruta al archivo HTML")
    parser.add_argument("-m", "--message", default=None, help="Mensaje del commit")
    parser.add_argument("--dry-run", action="store_true", help="Preview sin escribir ni subir")
    parser.add_argument("--update-only", action="store_true", help="Solo actualizar HTML, no subir a GitHub")
    parser.add_argument("--push-only", action="store_true", help="Solo subir a GitHub, no tocar HTML")
    parser.add_argument("--skip-market", action="store_true", help="No descargar datos de mercado (FRED/Banxico)")
    args = parser.parse_args()

    now = datetime.now().strftime("%d %b %Y %H:%M")
    commit_msg = args.message or f"Dashboard update — {now}"

    print("═" * 60)
    print("  VEMO Treasury · Weekly Investment Dashboard")
    print("═" * 60)

    # ── Paso 1: Actualizar HTML ─────────────────────────────────────
    if not args.push_only:
        if not os.path.exists(args.excel):
            print(f"\nError: No se encontró '{os.path.basename(args.excel)}'")
            sys.exit(1)
        if not os.path.exists(args.html):
            print(f"\nError: No se encontró '{os.path.basename(args.html)}'")
            sys.exit(1)

        print(f"\n── Extrayendo: {os.path.basename(args.excel)} ──")
        data = extract(args.excel)
        kpi = build_kpi_values(data)
        print(f"  {len(data['entities'])} entidades · Total H1: ${data['total_interest_h1']:,.0f} · YTD: {kpi['yield_ytd']}")

        print_report(data)

        print(f"\n── Actualizando: {os.path.basename(args.html)} ──")
        updated_html = update_html(args.html, data)

        # ── Paso 1.5: Datos de mercado (APIs) ─────────────────────────
        if not args.skip_market:
            print(f"\n── Datos de mercado (FRED + Banxico) ──")
            try:
                market = fetch_market_data()
                if market:
                    updated_html = update_market_html(updated_html, market)
                    print("  ✓ Gráficas actualizadas con datos reales")
                else:
                    print("  Usando datos existentes en el HTML")
            except Exception as e:
                print(f"  ⚠ Error: {e}")
                print("  Usando datos existentes en el HTML")
        else:
            print("\n  [skip-market] Datos de mercado omitidos")

        if args.dry_run:
            print("  [DRY RUN] No se escribieron cambios.")
        else:
            backup = args.html + ".bak"
            with open(args.html, "r", encoding="utf-8") as f:
                with open(backup, "w", encoding="utf-8") as b:
                    b.write(f.read())
            with open(args.html, "w", encoding="utf-8") as f:
                f.write(updated_html)
            print(f"  ✓ HTML actualizado (backup: {os.path.basename(backup)})")

    # ── Paso 2: Push a GitHub ───────────────────────────────────────
    if not args.update_only:
        push_to_github(commit_msg, dry_run=args.dry_run)

    print("\n✓ Listo.")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        pass
    except Exception as e:
        print(f"\n✗ Error inesperado: {e}")
    finally:
        input("\nPresiona Enter para cerrar...")
